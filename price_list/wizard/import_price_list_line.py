import base64
import io
import csv
from odoo.exceptions import UserError
from odoo import models, fields, _
import xlsxwriter
import openpyxl
import logging
_logger = logging.getLogger(__name__)


class ImportPriceListWizard(models.TransientModel):
    _name = 'import.price.list.wizard'
    _description = 'Import Price List Wizard'

    file = fields.Binary('File', required=True)
    file_name = fields.Char('File Name')
    import_to = fields.Selection([
        ('local', 'Local'),
        ('international_import', 'International/Import'),
        ('international_export', 'International/Export')],
        string='Import To', required=True)
    price_list_id = fields.Many2one('custom.price.list', string='Price List', required=True)

    def import_file(self):
        if not self.file:
            raise UserError(_("Please upload a file to import."))

        # Decode the uploaded Excel file
        file_content = base64.b64decode(self.file)
        file_data = io.BytesIO(file_content)
        workbook = openpyxl.load_workbook(file_data, read_only=True)
        sheet = workbook.active

        # Get the headers
        headers = [cell.value.lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        _logger.info("Headers: %s", headers)

        # Define header mappings for both formats
        header_mapping = {
            'product': 'product',
            'weight': 'weight',
            'state_id': 'state_id',
            'zone': 'zone',
            'sale_price': 'sale_price',
            'additional_services': 'additional_services',
            'taxes': 'taxes',
            'extra_weight': 'extra_weight',
            'extra_weight_sale_price': 'extra_weight_sale_price',
            'country_id': 'country_id',
            'custom_zone': 'custom_zone',
            'cost_price': 'cost_price',
            'extra_weight_cost_price': 'extra_weight_cost_price'
        }

        # Determine required fields based on import type
        if self.import_to == 'local':
            required_fields = ['product', 'weight', 'state_id', 'zone', 'sale_price', 'additional_services', 'taxes',
                               'extra_weight', 'extra_weight_sale_price']
            price_list_model = self.env['custom.price.list.local.line']
        else:
            required_fields = ['product', 'weight', 'country_id', 'zone', 'custom_zone', 'sale_price', 'cost_price',
                               'additional_services', 'taxes', 'extra_weight', 'extra_weight_sale_price',
                               'extra_weight_cost_price']
            price_list_model = self.env[
                'custom.price.list.international.import.line'] if self.import_to == 'international_import' \
                else self.env['custom.price.list.international.export.line']

        # Check if the Excel file contains all the required fields
        for field in required_fields:
            if field not in headers and field in header_mapping:
                if header_mapping[field] not in headers:
                    raise UserError(_("The file must contain the following columns: %s" % ', '.join(required_fields)))

        # Get the column index for each required field
        column_indices = {field: headers.index(field) for field in required_fields if field in headers}

        # Process the data row by row
        for row in sheet.iter_rows(min_row=2):
            row_data = {}
            for header, index in column_indices.items():
                row_data[header] = row[index].value

            _logger.info("Row Data: %s", row_data)

            # Find the product by name
            product = self.env['product.template'].search([('name', '=', row_data['product'])], limit=1)
            if not product:
                _logger.error("Product '%s' not found", row_data['product'])
                raise UserError(_("Product '%s' not found" % row_data['product']))

            # Split and search for additional services and taxes
            service_names = [name.strip() for name in row_data['additional_services'].split(',')] if row_data[
                'additional_services'] else []
            additional_services = self.env['account.tax'].search([('name', 'in', service_names)])

            tax_names = [name.strip() for name in row_data['taxes'].split(',')] if row_data['taxes'] else []
            taxes = self.env['account.tax'].search([('name', 'in', tax_names)])

            # Validate services and taxes
            if len(additional_services) != len(service_names):
                missing_services = set(service_names) - set(additional_services.mapped('name'))
                _logger.error("Missing additional services: %s", missing_services)
                raise UserError(_("The following additional services were not found: %s" % ', '.join(missing_services)))

            if len(taxes) != len(tax_names):
                missing_taxes = set(tax_names) - set(taxes.mapped('name'))
                _logger.error("Missing taxes: %s", missing_taxes)
                raise UserError(_("The following taxes were not found: %s" % ', '.join(missing_taxes)))

            # Prepare the values for creating a price list line
            line_values = {
                'product': product.id,
                'weight': float(row_data['weight']),
                'zone': row_data['zone'],
                'sale_price': float(row_data['sale_price']),
                'price_list_id': self.price_list_id.id,
                'additional_services': [(6, 0, additional_services.ids)],
                'taxes': [(6, 0, taxes.ids)],
                'extra_weight': float(row_data['extra_weight']),
                'extra_weight_sale_price': float(row_data['extra_weight_sale_price']),
            }

            # Handle country or state information based on import type
            if self.import_to == 'local':
                state = self.env['res.country.state'].search([('name', '=', row_data['state_id'])], limit=1)
                if not state:
                    _logger.error("State '%s' not found", row_data['state_id'])
                    raise UserError(_("State '%s' not found" % row_data['state_id']))
                line_values['state_id'] = state.id
            else:
                country = self.env['res.country'].search([('name', '=', row_data['country_id'])], limit=1)
                if not country:
                    _logger.error("Country '%s' not found", row_data['country_id'])
                    raise UserError(_("Country '%s' not found" % row_data['country_id']))
                line_values['country_id'] = country.id
                line_values['cost_price'] = float(row_data['cost_price'])
                line_values['custom_zone'] = row_data['custom_zone']
                line_values['extra_weight_cost_price'] = float(row_data['extra_weight_cost_price'])

            # Create the price list line
            price_list_model.create(line_values)

        return {'type': 'ir.actions.act_window_close'}

class ExportPriceListWizard(models.TransientModel):
    _name = 'export.price.list.wizard'
    _description = 'Export Price List Wizard'

    file_data = fields.Binary('File', readonly=True)
    file_name = fields.Char('File Name', readonly=True)
    export_from = fields.Selection([
        ('local', 'Local'),
        ('international_import', 'International/Import'),
        ('international_export', 'International/Export')],
        string='Export From', required=True)

    def export_file(self):
        active_id = self.env.context.get('active_id')
        if not active_id:
            raise UserError(_("No active price list to export."))

        # Determine the correct price list lines model based on export type
        if self.export_from == 'local':
            price_list_lines = self.env['custom.price.list.local.line'].search([('price_list_id', '=', active_id)])
        elif self.export_from == 'international_import':
            price_list_lines = self.env['custom.price.list.international.import.line'].search([('price_list_id', '=', active_id)])
        elif self.export_from == 'international_export':
            price_list_lines = self.env['custom.price.list.international.export.line'].search([('price_list_id', '=', active_id)])

        # Create an Excel file in-memory
        file_data = io.BytesIO()
        workbook = xlsxwriter.Workbook(file_data)
        worksheet = workbook.add_worksheet('Price List')

        # Define headers based on the type of export
        if self.export_from == 'local':
            header = ['product', 'weight', 'state_id', 'zone', 'sale_price', 'additional_services', 'taxes',
                      'extra_weight', 'extra_weight_sale_price']
        else:
            header = ['product', 'weight', 'country_id', 'zone', 'custom_zone', 'sale_price', 'cost_price',
                      'additional_services', 'taxes', 'extra_weight', 'extra_weight_sale_price',
                      'extra_weight_cost_price']

        # Write the headers
        for col_num, col_name in enumerate(header):
            worksheet.write(0, col_num, col_name)

        # Write the data
        for row_num, line in enumerate(price_list_lines, start=1):
            additional_services = ', '.join([service.name for service in line.additional_services])
            taxes = ', '.join([tax.name for tax in line.taxes])

            if self.export_from == 'local':
                worksheet.write(row_num, 0, line.product.name)
                worksheet.write(row_num, 1, line.weight)
                worksheet.write(row_num, 2, line.state_id.name)
                worksheet.write(row_num, 3, line.zone)
                worksheet.write(row_num, 4, line.sale_price)
                worksheet.write(row_num, 5, additional_services)
                worksheet.write(row_num, 6, taxes)
                worksheet.write(row_num, 7, line.extra_weight)
                worksheet.write(row_num, 8, line.extra_weight_sale_price)
            else:
                worksheet.write(row_num, 0, line.product.name)
                worksheet.write(row_num, 1, line.weight)
                worksheet.write(row_num, 2, line.country_id.name)
                worksheet.write(row_num, 3, line.zone)
                worksheet.write(row_num, 4, line.custom_zone)
                worksheet.write(row_num, 5, line.sale_price)
                worksheet.write(row_num, 6, line.cost_price)
                worksheet.write(row_num, 7, additional_services)
                worksheet.write(row_num, 8, taxes)
                worksheet.write(row_num, 9, line.extra_weight)
                worksheet.write(row_num, 10, line.extra_weight_sale_price)
                worksheet.write(row_num, 11, line.extra_weight_cost_price)

        # Save and close the workbook
        workbook.close()
        file_data.seek(0)

        # Prepare the binary file for download
        self.file_data = base64.b64encode(file_data.read())
        self.file_name = 'Price_List_{}.xlsx'.format(fields.Datetime.now().strftime('%Y-%m-%d %H-%M-%S'))

        return {
            'type': 'ir.actions.act_url',
            'url': '/web/content/?model=export.price.list.wizard&id=%s&field=file_data&download=true&filename=%s' % (
            self.id, self.file_name),
            'target': 'self',
        }
